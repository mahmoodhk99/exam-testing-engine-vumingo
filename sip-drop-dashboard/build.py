#!/usr/bin/env python3
"""Build the SIP Drop Dashboard from the two source reports.

Usage:
    pip install openpyxl xlrd
    python build.py contactdetail_1.xlsx Test_PAS.xls

Produces (next to this script):
    data.json    per-agent / per-session / per-job data
    index.html   self-contained dashboard (data embedded)

What it shows
-------------
For each agent it compares two independent measures of dropped nailer calls:

  * In Job Break Count  -- from the POM Agent Summary Report (Test_PAS.xls).
    When the sip:100@zain.com nailer call to a nailed agent drops during a
    job, POM records one "In Job Break" for that job.
  * sip:100 nailer drops -- number of sip:100@zain.com contacts that ended
    in Far End Disconnect for that agent, from the contact detail report.

The PAS report is hierarchical: Agent -> Session (Login/Logout) ->
Job (Job Attach / Job Detach) with a Call Count, Talk Duration and
In Job Break Count per job. A job whose In Job Break Count is greater
than one (the nailer dropped more than once within the same Job
Attach-Detach window) is flagged red. All durations are reported in
minutes.
"""
import json
import os
import re
import sys
from collections import Counter, defaultdict
from datetime import datetime, timedelta

import openpyxl
import xlrd

HERE = os.path.dirname(os.path.abspath(__file__))
TEMPLATE = os.path.join(HERE, "template.html")

# Column positions in the PAS "Agent Summary Report" (0-based).
PAS = dict(agent_id=0, agent_name=2, session_id=6, login=8, logout=10,
           campaign=13, task=17, attach=19, detach=21, call_count=23,
           talk=25, in_job_break=31, in_job_break_dur=33)

# Minimum off-gap (seconds) after a session for a "dropped then logged off"
# flag: the nailer dropped in the session and the agent was then off at least
# this long before the next login.
MIN_GAP = 60


def _dur_to_sec(s):
    s = str(s).strip()
    if not s or s == "None":
        return 0
    parts = s.split(":")
    try:
        if len(parts) == 3:
            return int(parts[0]) * 3600 + int(parts[1]) * 60 + int(parts[2])
        if len(parts) == 2:
            return int(parts[0]) * 60 + int(parts[1])
    except ValueError:
        return 0
    return 0


def _int(s):
    s = str(s).strip()
    try:
        return int(float(s))
    except ValueError:
        return 0


def _dt(s):
    """Parse a report timestamp; returns a datetime or None."""
    s = str(s).strip()
    for fmt in ("%m/%d/%y %I:%M:%S %p", "%m/%d/%Y %I:%M:%S %p"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            pass
    return None


def load_nailer_drops(detail_path):
    """sip:100@zain.com Far-End-Disconnect drops per agent extension.

    Returns (counts, events, period) where events maps an agent extension to
    a list of {"start", "end"} datetimes for each nailer call. "start" is the
    call start; "end" (= start + duration) is when the nailer actually
    dropped. start is used to attribute the drop to a login session; end is
    used to see whether the drop happened right before the agent logged off.
    """
    wb = openpyxl.load_workbook(detail_path, read_only=True, data_only=True)
    ws = wb["Details"]
    rows = list(ws.iter_rows(values_only=True))
    hdr = [str(h).strip() for h in rows[0]]
    dest = hdr.index("Destination #")
    start = hdr.index("Start Time")
    dur = hdr.index("Duration")
    drops = Counter()
    events = defaultdict(list)
    for r in rows[1:]:
        ext = re.sub(r"sip:(\d+)@.*", r"\1", str(r[dest]))
        st = _dt(r[start])
        try:
            seconds = int(float(r[dur]))
        except (TypeError, ValueError):
            seconds = 0
        end = st + timedelta(seconds=seconds) if st else None
        drops[ext] += 1
        events[ext].append({"start": st, "end": end})
    for ext in events:
        events[ext].sort(key=lambda e: (e["start"] is None, e["start"]))
    period = ""
    if "Summary" in wb.sheetnames:
        srows = list(wb["Summary"].iter_rows(values_only=True))
        if srows and len(srows[0]) > 1:
            period = str(srows[0][1])
    return drops, events, period


def load_pas(pas_path):
    """Parse the hierarchical PAS report into per-agent session/job data."""
    ws = xlrd.open_workbook(pas_path).sheet_by_index(0)

    def cv(r, c):
        return str(ws.cell_value(r, c)).strip()

    agents = {}          # ext -> agent dict
    cur = sess = None
    cur_campaign = ""
    for r in range(7, ws.nrows):
        aid = cv(r, PAS["agent_id"])
        if re.fullmatch(r"\d+(\.0)?", aid):
            cur = aid.replace(".0", "")
            m = re.search(r"\((.*?)\)", cv(r, PAS["agent_name"]))
            name = (m.group(1) if m else cv(r, PAS["agent_name"])).strip()
            agents.setdefault(cur, {"name": name, "ext": cur, "sessions": [],
                                    "ib": 0, "talk": 0, "jobs": 0})
            sess = None
        sid = cv(r, PAS["session_id"])
        if sid:
            sess = {"sid": sid, "login": cv(r, PAS["login"]),
                    "logout": cv(r, PAS["logout"]), "jobs": [], "ib": 0}
            agents[cur]["sessions"].append(sess)
        camp = cv(r, PAS["campaign"])
        if camp:
            cur_campaign = re.sub(r"\s*\(.*\)", "", camp).strip()
        if cv(r, PAS["task"]) == "Call_100" and cv(r, PAS["attach"]):
            ib = _int(cv(r, PAS["in_job_break"]))
            job = {"camp": cur_campaign, "attach": cv(r, PAS["attach"]),
                   "detach": cv(r, PAS["detach"]),
                   "cc": _int(cv(r, PAS["call_count"])),
                   "talk": _dur_to_sec(cv(r, PAS["talk"])), "ib": ib,
                   "ibdur": _dur_to_sec(cv(r, PAS["in_job_break_dur"]))}
            if sess is None:
                sess = {"sid": "?", "login": "", "logout": "",
                        "jobs": [], "ib": 0}
                agents[cur]["sessions"].append(sess)
            sess["jobs"].append(job)
            sess["ib"] += ib
            agents[cur]["ib"] += ib
            agents[cur]["talk"] += job["talk"]
            agents[cur]["jobs"] += 1
    return agents


def main():
    if len(sys.argv) >= 3:
        detail_path, pas_path = sys.argv[1], sys.argv[2]
    else:
        detail_path, pas_path = "contactdetail_1.xlsx", "Test_PAS.xls"

    drops, events, period = load_nailer_drops(detail_path)
    agents = load_pas(pas_path)

    out = []
    for ext, a in agents.items():
        a["sessions"] = [s for s in a["sessions"] if s["jobs"]]
        # attribute each sip:100 nailer drop to the login session whose
        # Login-Logout window contains the call start time (each drop to at
        # most one session, so session drop counts sum to the agent total).
        remaining = [e for e in events.get(ext, []) if e["start"] is not None]
        for s in a["sessions"]:
            lo, hi = _dt(s["login"]), _dt(s["logout"])
            if lo and hi:
                inside = [e for e in remaining if lo <= e["start"] <= hi]
                s["drops"] = len(inside)
                # latest moment the nailer actually dropped in this session
                ends = [e["end"] for e in inside if e["end"]]
                s["_lastDropEnd"] = max(ends) if ends else None
                remaining = [e for e in remaining
                             if not (lo <= e["start"] <= hi)]
            else:
                s["drops"] = 0
                s["_lastDropEnd"] = None
            # Abuse: a nailer drop that POM did NOT record as an In Job Break
            # means the agent dropped the nailer from the softphone. Any
            # excess of drops over in-job breaks in a session is abuse.
            s["abuse"] = max(0, s["drops"] - s["ib"])
        a["abuse"] = sum(s["abuse"] for s in a["sessions"])
        a["abuseSessions"] = sum(1 for s in a["sessions"] if s["abuse"] > 0)
        a["drops100"] = drops.get(ext, 0)
        # Order sessions latest-first and compute the gap between each session
        # and the next-earlier one: this session's login - previous logout.
        a["sessions"].sort(key=lambda s: (_dt(s["login"]) or datetime.min),
                           reverse=True)
        for i, s in enumerate(a["sessions"]):
            s["gap"] = None
            if i + 1 < len(a["sessions"]):
                hi = _dt(s["login"])
                lo = _dt(a["sessions"][i + 1]["logout"])
                if hi and lo:
                    s["gap"] = max(0, int((hi - lo).total_seconds()))
            # gapAfter = off time until the NEXT (later) session's login; that
            # later session sits at index i-1 and its gap is exactly this value.
            s["gapAfter"] = a["sessions"][i - 1]["gap"] if i > 0 else None
            # "Dropped then logged off": the nailer dropped in this session AND
            # the agent then went off before the next login (an unlogged gap).
            # dropLead = seconds from the last drop to logout (small => the
            # drop coincided with logging off).
            logout = _dt(s["logout"])
            lde = s.pop("_lastDropEnd", None)
            s["lastDrop"] = lde.strftime("%I:%M:%S %p") if lde else None
            s["dropLead"] = (max(0, int((logout - lde).total_seconds()))
                             if logout and lde and lde <= logout else None)
            # Strongest gap-abuse signal: an abuse drop (nailer hung up with no
            # In Job Break) in a session that is then followed by an unlogged
            # off-gap before the next login -- i.e. the agent dropped the
            # nailer and went off without it being logged.
            s["dropLogoff"] = bool(s["abuse"] > 0 and s["gapAfter"]
                                   and s["gapAfter"] >= MIN_GAP)
            # Order this session's jobs chronologically (by date/time) and
            # measure the "detached" idle time between each job's detach and
            # the next job's attach (plus login->first attach and last
            # detach->logout). This is the time the agent was not attached to
            # a job -- e.g. detach 06:07:07 PM, reattach 06:17:06 PM.
            login, logout = _dt(s["login"]), _dt(s["logout"])
            s["jobs"].sort(key=lambda j: (_dt(j["attach"]) or datetime.min))
            prev_end = login
            idle_total = 0
            for j in s["jobs"]:
                at, de = _dt(j["attach"]), _dt(j["detach"])
                j["idle"] = 0
                if at and prev_end:
                    j["idle"] = max(0, int((at - prev_end).total_seconds()))
                    idle_total += j["idle"]
                prev_end = de or prev_end
            s["endIdle"] = 0
            if logout and prev_end:
                s["endIdle"] = max(0, int((logout - prev_end).total_seconds()))
            idle_total += s["endIdle"]
            s["idle"] = idle_total
        a["dropLogoffSessions"] = sum(1 for s in a["sessions"]
                                      if s["dropLogoff"])
        out.append(a)

    data = {"period": period, "origin": "sip:100@zain.com",
            "totalDrops100": sum(drops.values()),
            "totalIB": sum(a["ib"] for a in out), "agents": out}
    json.dump(data, open(os.path.join(HERE, "data.json"), "w"),
              separators=(",", ":"))

    tpl = open(TEMPLATE).read()
    html = tpl.replace("__DATA__", json.dumps(data, separators=(",", ":")))
    open(os.path.join(HERE, "index.html"), "w").write(html)
    print("Built %d agents | In Job Break drops: %d | sip:100 drops: %d | "
          "abuse drops (no break): %d across %d agents" %
          (len(out), data["totalIB"], data["totalDrops100"],
           sum(a["abuse"] for a in out),
           sum(1 for a in out if a["abuse"] > 0)))


if __name__ == "__main__":
    main()
